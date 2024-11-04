import os
import lpips
import torch
from PIL import Image
from torchvision import transforms
import itertools
import pandas as pd
from openpyxl import Workbook, load_workbook

# compute LPIPS scores between images in a folder
def Lpips_compute(dataset_path, results_file):
    loss_fn = lpips.LPIPS(net='alex')

    transform = transforms.Compose([
        transforms.Resize((256, 256)),
        transforms.ToTensor()
    ])

    image_files = [os.path.join(dataset_path, f) for f in os.listdir(dataset_path) if f.endswith(('png', 'jpg', 'jpeg'))]

    results = []
    for img1_path, img2_path in itertools.combinations(image_files, 2):
        img1 = Image.open(img1_path).convert('RGB')
        img2 = Image.open(img2_path).convert('RGB')
        img1 = transform(img1).unsqueeze(0)
        img2 = transform(img2).unsqueeze(0)

        with torch.no_grad():
            distance = loss_fn(img1, img2)

        results.append([os.path.basename(img1_path), os.path.basename(img2_path), distance.item()])
        print(f"Computed LPIPS between {img1_path} and {img2_path}: {distance.item()}")

    df = pd.DataFrame(results, columns=["Image1", "Image2", "LPIPS_Distance"])
    return df

# cross-model LPIPS scores
def Lpips_cross_model(images_folder, output_file):
    model_folders = [f for f in os.listdir(images_folder) if os.path.isdir(os.path.join(images_folder, f))]
    
    # images with the same name from different models
    loss_fn = lpips.LPIPS(net='alex')
    transform = transforms.Compose([
        transforms.Resize((256, 256)),
        transforms.ToTensor()
    ])
    
    results = []
    
    for model1, model2 in itertools.combinations(model_folders, 2):
        model1_path = os.path.join(images_folder, model1)
        model2_path = os.path.join(images_folder, model2)
        
        model1_images = {os.path.basename(f): os.path.join(model1_path, f) for f in os.listdir(model1_path) if f.endswith(('png', 'jpg', 'jpeg'))}
        model2_images = {os.path.basename(f): os.path.join(model2_path, f) for f in os.listdir(model2_path) if f.endswith(('png', 'jpg', 'jpeg'))}
        
        common_images = set(model1_images.keys()).intersection(model2_images.keys())
        
        # LPIPS for common images
        for image_name in common_images:
            img1_path = model1_images[image_name]
            img2_path = model2_images[image_name]
            
            img1 = Image.open(img1_path).convert('RGB')
            img2 = Image.open(img2_path).convert('RGB')
            img1 = transform(img1).unsqueeze(0)
            img2 = transform(img2).unsqueeze(0)
            
            with torch.no_grad():
                distance = loss_fn(img1, img2)

            results.append([model1, model2, image_name, distance.item()])
            print(f"Computed LPIPS between {image_name} from {model1} and {model2}: {distance.item()}")
    
    df = pd.DataFrame(results, columns=["Model1", "Model2", "Image", "LPIPS_Distance"])
    return df

def save_to_excel(df, output_file, sheet_name):
    if os.path.exists(output_file):
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(output_file, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# calculate LPIPS scores for all models in Images folder
def main(images_folder, output_file, folder_to_calculate=None, cross_model=False):
    # If a specific folder is given,
    if folder_to_calculate:
        dataset_path = os.path.join(images_folder, folder_to_calculate)
        df = Lpips_compute(dataset_path, output_file)
        save_to_excel(df, output_file, sheet_name=folder_to_calculate)
    else:
        # loops over each folder in imgs directory
        for model_folder in os.listdir(images_folder):
            model_path = os.path.join(images_folder, model_folder)
            if os.path.isdir(model_path):
                df = Lpips_compute(model_path, output_file)
                save_to_excel(df, output_file, sheet_name=model_folder)

    # Compute the average LPIPS scores per model - saves to seperate sheet
    workbook = load_workbook(output_file)
    average_lpips_data = []
    for sheet in workbook.sheetnames:
        df = pd.read_excel(output_file, sheet_name=sheet)
        avg_lpips_score = df["LPIPS_Distance"].mean()
        average_lpips_data.append([sheet, avg_lpips_score])

    avg_df = pd.DataFrame(average_lpips_data, columns=["Model", "Average_LPIPS_Score"])
    save_to_excel(avg_df, output_file, sheet_name='Average_Scores')
    
    # Compute cross-model 
    if cross_model:
        cross_model_df = Lpips_cross_model(images_folder, output_file)
        save_to_excel(cross_model_df, output_file, sheet_name='Cross_Model_Comparisons')

if __name__ == "__main__":
    images_folder = os.path.expanduser("~/Desktop/Images")
    output_file = os.path.expanduser("~/Desktop/lpips_results.xlsx")
    folder_to_calculate = None  # Set to specific folder name 
    cross_model = False  # Set to True if cross-model LPIPS scores are to be calculated
    main(images_folder, output_file, folder_to_calculate, cross_model)
